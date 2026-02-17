from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from app.models.document import DocumentType, FundaePaymentType, PaymentMethod

REQUIRED_CLIENT_COLUMNS = {"full_name", "nif", "phone"}
SPANISH_MONTHS = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


@dataclass
class ImportedRow:
    data: dict[str, Any]
    row_number: int


@dataclass
class ImportResult:
    clients_created: int
    clients_updated: int
    documents_created: int
    errors: list[str]


class ImportValidationError(Exception):
    pass


class SpreadsheetImporter:
    SUPPORTED_SUFFIXES = {".csv", ".xlsx"}

    def import_file(
        self,
        file_path: str | Path,
        column_mapping: dict[str, str] | None = None,
    ) -> list[ImportedRow]:
        path = Path(file_path)
        if path.suffix.lower() not in self.SUPPORTED_SUFFIXES:
            raise ImportValidationError(f"Tipo de archivo no soportado: {path.suffix}")

        rows = self._read_csv(path) if path.suffix.lower() == ".csv" else self._read_xlsx(path)
        mapped = self._apply_mapping(rows, column_mapping or {})
        self._validate_required_columns(mapped)
        return mapped

    def _read_csv(self, path: Path) -> list[dict[str, Any]]:
        with path.open("r", newline="", encoding="utf-8-sig") as stream:
            reader = csv.DictReader(stream)
            return [dict(row) for row in reader]

    def _read_xlsx(self, path: Path) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportValidationError("openpyxl es obligatorio para importar archivos .xlsx") from exc

        workbook = load_workbook(path, data_only=True)
        special_rows = self._read_known_real_world_xlsx(path, workbook)
        if special_rows is not None:
            return special_rows

        sheet = workbook.active
        headers: list[str] = []
        records: list[dict[str, Any]] = []

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                headers = [str(value).strip() if value is not None else "" for value in row]
                continue

            data = {headers[col_idx]: value for col_idx, value in enumerate(row) if col_idx < len(headers) and headers[col_idx]}
            if any(v is not None and str(v).strip() != "" for v in data.values()):
                records.append(data)

        return records

    def _read_known_real_world_xlsx(self, path: Path, workbook: Any) -> list[dict[str, Any]] | None:
        sheet_names = {name.lower(): name for name in workbook.sheetnames}
        stem = path.stem.lower()

        if "cap" in stem and "hoja2" in sheet_names:
            return self._read_cap_real_sheet(path, workbook[sheet_names["hoja2"]])

        if "tarjetas" in stem and "conductores s" in sheet_names and "empresas" in sheet_names:
            return self._read_tarjetas_real_sheets(
                path,
                workbook[sheet_names["conductores s"]],
                workbook[sheet_names["empresas"]],
            )

        return None

    def _read_cap_real_sheet(self, path: Path, sheet: Any) -> list[dict[str, Any]]:
        year = _extract_year_from_filename(path)
        if year is None:
            raise ImportValidationError(
                "No se pudo determinar el año desde el nombre del fichero CAP (ejemplo esperado: 'CAP 2025.xlsx')."
            )

        records: list[dict[str, Any]] = []
        current_issue_date: date | None = None

        for row_idx in range(1, sheet.max_row + 1):
            row_values = [sheet.cell(row=row_idx, column=col).value for col in range(1, 11)]
            first_cell_title = _clean_text(row_values[0])
            block_title = _clean_text(row_values[1])
            dni_header = _normalize_header_token(row_values[3])
            curso_header = _normalize_header_token(row_values[4])
            empresa_header = _normalize_header_token(row_values[5])
            tlf_header = _normalize_header_token(row_values[6])

            first_cell_issue = _parse_cap_block_issue_date(first_cell_title, year)
            if first_cell_issue:
                current_issue_date = first_cell_issue
                # La primera fila de cada subtabla es un encabezado con fecha, nunca un cliente.
                continue

            if dni_header == "dni" and curso_header == "curso" and empresa_header == "empresa" and tlf_header == "tlf":
                parsed_issue = _parse_cap_block_issue_date(block_title, year)
                if parsed_issue:
                    current_issue_date = parsed_issue
                continue

            title_issue_date = _parse_cap_block_issue_date(block_title, year)
            if title_issue_date and _looks_like_cap_header_row(block_title, row_values[3], row_values[4], row_values[5], row_values[6]):
                current_issue_date = title_issue_date
                continue

            full_name = _clean_text(row_values[1])
            nif = _normalize_identifier(row_values[3])
            if not full_name and not nif:
                continue
            if _is_cap_fake_nif(nif):
                continue
            if not full_name or not nif:
                continue

            course_number = _normalize_course_number(row_values[4])
            company = _clean_text(row_values[5])
            phone = _normalize_cap_phone(row_values[6])
            fundae = _parse_cap_fundae(row_values[7])
            flag_c, flag_d = _parse_permiso_flags(row_values[8])

            expiry_date = _add_years_safe(current_issue_date, 5) if current_issue_date else None

            records.append(
                {
                    "full_name": full_name,
                    "nif": nif,
                    "phone": phone,
                    "company": company,
                    "document_type": "cap",
                    "issue_date": current_issue_date,
                    "expiry_date": expiry_date,
                    "course_number": course_number,
                    "renewed_with_us": True,
                    "payment_method": "empresa",
                    "fundae": fundae,
                }
            )

            if flag_c or flag_d:
                records.append(
                    {
                        "full_name": full_name,
                        "nif": nif,
                        "phone": phone,
                        "company": company,
                        "document_type": "driving_license",
                        "flag_permiso_c": flag_c,
                        "flag_permiso_d": flag_d,
                    }
                )

        return records

    def _read_tarjetas_real_sheets(self, path: Path, conductores_sheet: Any, empresas_sheet: Any) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        year = _extract_year_from_filename(path) or date.today().year

        for row in self._iter_tarjetas_rows(conductores_sheet):
            full_name = _clean_text(row.get("nombre"))
            nif = _normalize_identifier(row.get("dni/cif"))
            if not full_name or not nif:
                continue

            issue_date = _coerce_date(row.get("fecha registro"))
            expiry_date = _add_years_safe(issue_date, 5) if issue_date else None

            records.append(
                {
                    "full_name": full_name,
                    "nif": nif,
                    "phone": _normalize_generic_phone(row.get("telefono")),
                    "document_type": "tachograph_card",
                    "issue_date": issue_date,
                    "expiry_date": expiry_date,
                }
            )

        for row in self._iter_tarjetas_rows(empresas_sheet):
            full_name = _clean_text(row.get("nombre"))
            nif = _normalize_identifier(row.get("dni/cif"))
            if not full_name or not nif:
                continue

            issue_date = _coerce_date(row.get("fecha registro"))
            expiry_date = _add_years_safe(issue_date, 5) if issue_date else None

            records.append(
                {
                    "full_name": full_name,
                    "nif": nif,
                    "phone": _normalize_generic_phone(row.get("telefono")),
                    "company": full_name,
                    "document_type": "tachograph_card",
                    "issue_date": issue_date,
                    "expiry_date": expiry_date,
                }
            )

            apodera_raw = row.get("apodera")
            if _has_meaningful_value(apodera_raw):
                apodera_date = _coerce_date(apodera_raw)
                poa_expiry = expiry_date or _add_years_safe(apodera_date, 5) or date(year + 5, 1, 1)
                records.append(
                    {
                        "full_name": full_name,
                        "nif": nif,
                        "phone": _normalize_generic_phone(row.get("telefono")),
                        "company": full_name,
                        "document_type": "power_of_attorney",
                        "flag_fran": True,
                        "expiry_fran": poa_expiry,
                    }
                )

        return records

    def _iter_tarjetas_rows(self, sheet: Any) -> list[dict[str, Any]]:
        header_values = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
        normalized_headers = [_normalize_header_token(value) for value in header_values]
        index_map = {key: idx for idx, key in enumerate(normalized_headers) if key}

        out: list[dict[str, Any]] = []
        for row_idx in range(2, sheet.max_row + 1):
            row_values = [sheet.cell(row=row_idx, column=col).value for col in range(1, sheet.max_column + 1)]
            if not any(_has_meaningful_value(value) for value in row_values):
                continue

            out.append(
                {
                    "nombre": _pick_by_headers(row_values, index_map, ["nombre"]),
                    "dni/cif": _pick_by_headers(row_values, index_map, ["dni/cif", "dni cif", "dni", "cif"]),
                    "telefono": _pick_by_headers(row_values, index_map, ["telefono", "tlf"]),
                    "fecha registro": _pick_by_headers(row_values, index_map, ["fecha registro"]),
                    "apodera": _pick_by_headers(row_values, index_map, ["apodera"]),
                }
            )
        return out

    def _apply_mapping(self, rows: list[dict[str, Any]], column_mapping: dict[str, str]) -> list[ImportedRow]:
        imported: list[ImportedRow] = []
        for idx, row in enumerate(rows, start=2):
            mapped: dict[str, Any] = {}
            for source_column, value in row.items():
                target_key = column_mapping.get(source_column, source_column)
                mapped[target_key] = self._normalize_value(value)

            imported.append(ImportedRow(data=mapped, row_number=idx))

        return imported

    def _validate_required_columns(self, rows: list[ImportedRow]) -> None:
        if not rows:
            raise ImportValidationError("El archivo no contiene filas de datos.")

        keys = set(rows[0].data.keys())
        missing = REQUIRED_CLIENT_COLUMNS - keys
        if missing:
            raise ImportValidationError(f"Faltan columnas obligatorias: {', '.join(sorted(missing))}")

    def _normalize_value(self, value: Any) -> Any:
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
            return value
        return value


def parse_document_type(value: Any) -> DocumentType | None:
    if not value:
        return None

    token = str(value).strip().lower()
    aliases = {
        "dni": DocumentType.DNI,
        "carnet": DocumentType.DRIVING_LICENSE,
        "carnet_conducir": DocumentType.DRIVING_LICENSE,
        "permiso_conducir": DocumentType.DRIVING_LICENSE,
        "driving_license": DocumentType.DRIVING_LICENSE,
        "cap": DocumentType.CAP,
        "tachograph": DocumentType.TACHOGRAPH_CARD,
        "tacografo": DocumentType.TACHOGRAPH_CARD,
        "tarjeta_tacografo": DocumentType.TACHOGRAPH_CARD,
        "tachograph_card": DocumentType.TACHOGRAPH_CARD,
        "poder_notarial": DocumentType.POWER_OF_ATTORNEY,
        "power_of_attorney": DocumentType.POWER_OF_ATTORNEY,
        "power of attorney": DocumentType.POWER_OF_ATTORNEY,
        "otro": DocumentType.OTHER,
        "other": DocumentType.OTHER,
    }
    return aliases.get(token)


def parse_payment_method(value: Any) -> PaymentMethod | None:
    if not value:
        return None
    token = str(value).strip().lower()
    aliases = {
        "efectivo": PaymentMethod.EFECTIVO,
        "cash": PaymentMethod.EFECTIVO,
        "visa": PaymentMethod.VISA,
        "empresa": PaymentMethod.EMPRESA,
        "company": PaymentMethod.EMPRESA,
        "fundae": PaymentMethod.EMPRESA,
    }
    return aliases.get(token)


def parse_fundae_payment_type(value: Any) -> FundaePaymentType | None:
    if not value:
        return None
    token = str(value).strip().lower()
    aliases = {
        "recibo": FundaePaymentType.RECIBO,
        "receipt": FundaePaymentType.RECIBO,
        "transferencia": FundaePaymentType.TRANSFERENCIA,
        "transfer": FundaePaymentType.TRANSFERENCIA,
    }
    return aliases.get(token)


def to_date(value: Any) -> date | None:
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        token = value.strip().lower()
        return token in {"1", "true", "yes", "y", "si", "s", "verdadero"}
    return False


def _normalize_header_token(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("/", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _extract_year_from_filename(path: Path) -> int | None:
    match = re.search(r"(19|20)\d{2}", path.stem)
    if not match:
        return None
    return int(match.group(0))


def _parse_cap_block_issue_date(title: str | None, year: int) -> date | None:
    if not title:
        return None

    normalized = _normalize_header_token(title)
    month: int | None = None
    for month_name, month_number in SPANISH_MONTHS.items():
        if month_name in normalized:
            month = month_number
            break
    if month is None:
        return None

    day_match = re.search(r"\b(\d{1,2})\b", str(title))
    if not day_match:
        return None
    day = int(day_match.group(1))
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return " ".join(text.split())


def _normalize_identifier(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            value = int(value)
    text = str(value).strip().upper().replace(" ", "")
    return text


def _normalize_course_number(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            value = int(value)
    text = str(value).strip()
    return text or None


def _normalize_cap_phone(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value)
    text = str(value).strip().replace(" ", "")
    return text if text.isdigit() else ""


def _normalize_generic_phone(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value)
    text = str(value).strip()
    return text


def _parse_cap_fundae(value: Any) -> bool:
    if value is None:
        return False
    text = _clean_text(value)
    if not text:
        return False
    return _normalize_header_token(text) != "no"


def _parse_permiso_flags(value: Any) -> tuple[bool, bool]:
    if value is None:
        return False, False
    token = _normalize_header_token(value).replace(" ", "")
    return "c" in token, "d" in token


def _coerce_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return to_date(value)


def _add_years_safe(value: date | None, years: int) -> date | None:
    if value is None:
        return None
    try:
        return value.replace(year=value.year + years)
    except ValueError:
        # 29-feb -> 28-feb on non-leap years.
        return value.replace(month=2, day=28, year=value.year + years)


def _has_meaningful_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _pick_by_headers(row_values: list[Any], index_map: dict[str, int], candidates: list[str]) -> Any:
    for candidate in candidates:
        idx = index_map.get(_normalize_header_token(candidate))
        if idx is None or idx >= len(row_values):
            continue
        return row_values[idx]
    return None


def _is_cap_fake_nif(nif: str) -> bool:
    token = _normalize_header_token(nif)
    return token in {"dni", "dni cif", "cif", "nombre"}


def _looks_like_cap_header_row(title: str | None, dni_cell: Any, curso_cell: Any, empresa_cell: Any, tlf_cell: Any) -> bool:
    if not title:
        return False

    normalized_title = _normalize_header_token(title)
    dni_token = _normalize_header_token(dni_cell)
    curso_token = _normalize_header_token(curso_cell)
    empresa_token = _normalize_header_token(empresa_cell)
    tlf_token = _normalize_header_token(tlf_cell)

    has_dni_hint = "dni" in normalized_title or dni_token in {"dni", "dni cif", "cif"}
    has_header_hints = any(
        token in {"curso", "empresa", "tlf"}
        for token in (curso_token, empresa_token, tlf_token)
    )
    return has_dni_hint or has_header_hints
